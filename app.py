import openpyxl

def createbill():
    nm = input("Enter Customer Name: ")
    cnt = input("Enter Contact: ")

    items = []  # store all items

    for i in range(1, 5):
        item = input(f"Enter Item {i}: ")
        quantity = int(input("Enter Quantity: "))
        price = float(input("Enter Per Product Price: "))
        total = quantity * price
        items.append([item, quantity, price, total])

    # Calculate grand total
    grand_total = sum(row[3] for row in items)

    # Save to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill"

    # Customer details
    ws["A1"] = "Customer Name"
    ws["B1"] = nm
    ws["A2"] = "Contact"
    ws["B2"] = cnt

    # Table header
    ws.append([])
    ws.append(["Item", "Quantity", "Price", "Total"])

    # Add items
    for row in items:
        ws.append(row)

    # Grand total
    ws.append(["", "", "Grand Total", grand_total])

    # Save file
    filename = f"Bill_{nm}.xlsx"
    wb.save(filename)

    print(f"\n✅ Bill saved as {filename}")
    menu()


def menu():
    print("\n=== Welcome to My Billing System ===")
    print("1. Create New Bill")
    print("2. Exit")
    op = input("Enter Option: ")

    if op == "1":
        createbill()
    elif op == "2":
        print("Goodbye!")
        exit()
    else:
        print("Invalid choice!")
        menu()


menu()
